"""Automated Daily CAPI Sync Engine: Downloads latest Google Sheet 'DATHEN', detects NEW leads, and synchronizes with Meta CAPI Pixel 902489598915870."""

import os
import sys
import json
import openpyxl
import hashlib
import unicodedata
import requests
from pathlib import Path
from datetime import datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

MAIN_PIXEL_ID = "902489598915870"
SHEET_ID = "1zq0nnHqKgtsZBZnEKknM55qI_wjnm_Z5MzPBDLBD1jc"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

def load_env():
    env_path = Path(".env")
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ[k.strip()] = v.strip()

def strip_accents(text: str) -> str:
    if not text: return ""
    text = unicodedata.normalize('NFD', str(text))
    return ''.join(c for c in text if unicodedata.category(c) != 'Mn').lower().strip()

def normalize_phone_vn(phone: str) -> str:
    digits = "".join(filter(str.isdigit, str(phone)))
    if not digits: return ""
    if digits.startswith("0") and len(digits) >= 9:
        digits = "84" + digits[1:]
    elif not digits.startswith("84") and len(digits) == 9:
        digits = "84" + digits
    return digits

def hash_sha256(value: str) -> str:
    if not value: return ""
    return hashlib.sha256(str(value).encode("utf-8")).hexdigest()

def split_vietnamese_name(full_name: str):
    cleaned = strip_accents(full_name)
    parts = cleaned.split()
    if not parts: return "", ""
    if len(parts) == 1: return "", parts[0]
    return parts[0], parts[-1]

def map_branch_city(branch_name: str) -> str:
    b = strip_accents(branch_name).upper()
    if "DA LAT" in b: return "dalat"
    if "BIEN HOA" in b or "GIA KIEM" in b: return "bienhoa"
    if "CAN THO" in b or "THOT NOT" in b: return "cantho"
    if "VUNG TAU" in b or "BA RIA" in b or "PHUOC TINH" in b or "XUYEN MOC" in b: return "vungtau"
    if "BINH DUONG" in b or "DI AN" in b or "MINH HOA" in b: return "thuylai"
    if "TAY NINH" in b: return "tayninh"
    if "QUY NHON" in b: return "quynhon"
    if "DA NANG" in b: return "danang"
    if "CA MAU" in b: return "camau"
    if "BAC LIEU" in b: return "baclieu"
    if "SOC TRANG" in b: return "soctrang"
    if "DONG THAP" in b: return "dongthap"
    if "HOA BINH" in b: return "hoabinh"
    return "hochiminh"

def run_daily_auto_sync():
    load_env()
    token = os.environ.get("META_API_TOKEN")
    pixel_id = os.environ.get("META_PIXEL_ID", MAIN_PIXEL_ID)

    print(f"\n=========================================================================================")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] DAILY AUTOMATED CAPI SYNC STARTED")
    print(f"=========================================================================================")

    # 1. Download Latest Google Sheet
    xlsx_path = Path("data/google_leads_latest.xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"[FETCH] Downloading latest Google Sheet workbook from Google Drive...")
    try:
        res = requests.get(XLSX_URL, timeout=30)
        if res.status_code == 200 and len(res.content) > 1000:
            xlsx_path.write_bytes(res.content)
            print(f"[FETCH SUCCESS] Downloaded latest workbook ({len(res.content):,} bytes).")
        else:
            print(f"[FETCH ERROR] HTTP {res.status_code}")
            return
    except Exception as exc:
        print(f"[FETCH FAILED] {exc}")
        return

    # 2. Load Synced Tracker Ledger
    tracker_path = Path(".claude-ads/manifests/synced_leads_tracker.json")
    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    synced_set = set()
    if tracker_path.exists():
        try:
            tracker_data = json.loads(tracker_path.read_text(encoding="utf-8"))
            synced_set = set(tracker_data.get("synced_event_ids", []))
        except Exception:
            pass

    print(f"[TRACKER] Currently tracking {len(synced_set):,} previously synchronized event IDs.")

    # 3. Parse Sheet DATHEN
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    dathen_sheet_name = None
    for name in wb.sheetnames:
        if "DATHEN" in name.upper() or "ĐÃ HẸN" in name.upper():
            dathen_sheet_name = name
            break
    if not dathen_sheet_name: dathen_sheet_name = wb.sheetnames[0]

    sheet = wb[dathen_sheet_name]
    raw_rows = list(sheet.iter_rows(values_only=True))

    col_date, col_name, col_phone, col_source, col_branch, col_service = 0, 1, 2, 3, 4, 5

    if len(raw_rows) >= 3:
        h_row = [str(c).upper().strip() if c else "" for c in raw_rows[2]]
        for idx, c in enumerate(h_row):
            if "NGÀY" in c: col_date = idx
            elif "HỌ TÊN" in c or "HO" in c: col_name = idx
            elif "SĐT" in c or "SDT" in c: col_phone = idx
            elif "NGUỒN" in c: col_source = idx
            elif "CHI NHÁNH" in c: col_branch = idx
            elif "DV" in c or "DỊCH VỤ" in c: col_service = idx

    new_leads_to_sync = []
    
    for row in raw_rows[3:]:
        if not row or len(row) <= col_phone: continue
        phone = str(row[col_phone]).strip() if row[col_phone] is not None else ""
        name = str(row[col_name]).strip() if col_name < len(row) and row[col_name] is not None else ""

        if not phone or phone.upper() in ["NONE", "SĐT", "SDT", "N/A", ""]: continue

        ph_norm = normalize_phone_vn(phone)
        date_str = str(row[col_date]).strip()[:10] if col_date < len(row) and row[col_date] is not None else "2026-06-15"
        event_id = f"LEAD_DATHEN_{ph_norm}_{date_str}"

        # Deduplication check against local tracker
        if event_id not in synced_set:
            source = str(row[col_source]).strip().upper() if col_source < len(row) and row[col_source] is not None else "FACEBOOK"
            branch = str(row[col_branch]).strip().upper() if col_branch < len(row) and row[col_branch] is not None else "HCM"
            service = str(row[col_service]).strip() if col_service < len(row) and row[col_service] is not None else "Nha Khoa"

            new_leads_to_sync.append({
                "event_id": event_id,
                "date": date_str,
                "name": name,
                "phone": phone,
                "ph_norm": ph_norm,
                "source": source,
                "branch": branch,
                "service": service
            })

    total_new = len(new_leads_to_sync)
    print(f"[DELTA DETECTION] Found {total_new:,} NEW Lead events to sync to Meta CAPI.")

    if total_new == 0:
        print("[AUTO SYNC] All leads up to date! Zero new leads detected.")
        return

    # 4. Push New Leads to Meta CAPI
    url = f"https://graph.facebook.com/v20.0/{pixel_id}/events"
    batch_size = 100
    success_count = 0

    for i in range(0, total_new, batch_size):
        batch = new_leads_to_sync[i:i + batch_size]
        events_payload = []

        for lead in batch:
            ph_norm = lead["ph_norm"]
            ln, fn = split_vietnamese_name(lead["name"])
            city = map_branch_city(lead["branch"])
            svc = lead["service"]

            u_data = {"country": [hash_sha256("vn")]}
            if ph_norm: u_data["ph"] = [hash_sha256(ph_norm)]
            if fn: u_data["fn"] = [hash_sha256(fn)]
            if ln: u_data["ln"] = [hash_sha256(ln)]
            if city: u_data["ct"] = [hash_sha256(city)]
            if ph_norm: u_data["external_id"] = [hash_sha256(f"LEAD_{ph_norm}")]

            events_payload.append({
                "event_name": "Lead",
                "event_time": int(datetime.now().timestamp()),
                "event_id": lead["event_id"],
                "action_source": "system_generated",
                "user_data": u_data,
                "custom_data": {
                    "content_name": svc,
                    "content_category": svc,
                    "branch": lead["branch"],
                    "lead_type": "AppointmentBooked"
                }
            })

        payload = {"data": events_payload, "access_token": token}
        try:
            res = requests.post(url, json=payload, timeout=15).json()
            if "events_received" in res:
                rec = res["events_received"]
                success_count += rec
                for lead in batch:
                    synced_set.add(lead["event_id"])
        except Exception as exc:
            print(f"  -> Batch Error: {exc}")

    # 5. Save Updated Synced Tracker Ledger
    updated_tracker = {
        "last_sync_timestamp": datetime.now().isoformat(),
        "total_tracked_events": len(synced_set),
        "synced_event_ids": list(synced_set)
    }
    tracker_path.write_text(json.dumps(updated_tracker, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n=========================================================================================")
    print(f"[AUTO SYNC COMPLETE] Successfully pushed {success_count:,} NEW Leads to Meta Pixel {pixel_id}!")
    print(f"[TRACKER UPDATED] Total Tracked Event IDs: {len(synced_set):,}")
    print(f"=========================================================================================")

if __name__ == "__main__":
    run_daily_auto_sync()
