"""Debug column mappings and headers across all 15 Excel files in customer/2026."""

import sys
import openpyxl
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

def debug_files():
    customer_dir = Path("customer/2026")
    files = sorted(list(customer_dir.glob("*.xlsx")))

    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheet = None
        for sname in wb.sheetnames:
            if "TỔNG" in sname.upper() and "COPY" not in sname.upper():
                sheet = wb[sname]
                break
        if not sheet:
            sheet = wb[wb.sheetnames[0]]

        rows = list(sheet.iter_rows(values_only=True, max_row=10))
        wb.close()

        print(f"\n==========================================")
        print(f"FILE: {f.name}")
        print(f"==========================================")
        for r_idx, row in enumerate(rows, start=1):
            row_str = [str(c) if c is not None else "" for c in row[:12]]
            if any("HỌ TÊN" in c.upper() or "DATE" in c.upper() or "SĐT" in c.upper() for c in row_str):
                print(f"HEADER (Row {r_idx}): {row_str}")
            elif r_idx <= 4:
                print(f"Data (Row {r_idx}): {row_str}")

if __name__ == "__main__":
    debug_files()
