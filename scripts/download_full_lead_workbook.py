"""Download full Google Sheet XLSX workbook and parse sheet tab 'DATHEN' (19,000+ Booked Leads)."""

import sys
import openpyxl
import requests
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

SHEET_ID = "1zq0nnHqKgtsZBZnEKknM55qI_wjnm_Z5MzPBDLBD1jc"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

def fetch_and_inspect_dathen():
    print(f"[GOOGLE SHEET FULL WORKBOOK] Downloading XLSX from Google Sheet ID: {SHEET_ID}...")
    
    xlsx_path = Path("data/google_leads_full.xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        res = requests.get(XLSX_URL, timeout=30)
        if res.status_code == 200 and len(res.content) > 1000:
            xlsx_path.write_bytes(res.content)
            print(f"[SUCCESS] Downloaded full XLSX workbook ({len(res.content):,} bytes).")
        else:
            print(f"[ERROR] HTTP status {res.status_code}")
            return
    except Exception as exc:
        print(f"[ERROR] Download failed: {exc}")
        return

    # Inspect Workbook Sheet Names
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print("\nAvailable Sheet Names in Workbook:")
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"  - Sheet: '{name}' | Max Rows: {sheet.max_row:,}")

    target_sheet_name = None
    for name in wb.sheetnames:
        if "DATHEN" in name.upper() or "ĐÃ HẸN" in name.upper() or "DA HEN" in name.upper():
            target_sheet_name = name
            break

    if not target_sheet_name and wb.sheetnames:
        target_sheet_name = wb.sheetnames[0]

    print(f"\n[TARGET SHEET] Inspecting Sheet Tab: '{target_sheet_name}'...")
    target_sheet = wb[target_sheet_name]
    
    rows = list(target_sheet.iter_rows(values_only=True))
    print(f"Total Rows in Sheet '{target_sheet_name}': {len(rows):,}")

    if rows:
        print("\nHeader Row (Row 1 or 2):")
        for idx, val in enumerate(rows[0][:15]):
            print(f"  Col {idx:2d}: {val}")

        print("\nTop 5 Sample Data Rows:")
        for r_idx, row in enumerate(rows[1:6], start=1):
            print(f"Row {r_idx}: {row[:10]}")

if __name__ == "__main__":
    fetch_and_inspect_dathen()
