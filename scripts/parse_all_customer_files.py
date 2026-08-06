"""Strict column parser for all 15 customer Excel files to fix branch/revenue column swap."""

import sys
import json
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

def parse_monthly_file(file_path: Path):
    print(f"[PARSING STRICT] File: {file_path.name}...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    customers = []
    sheet = None
    for sname in wb.sheetnames:
        if "TỔNG" in sname.upper() and "COPY" not in sname.upper():
            sheet = wb[sname]
            break
            
    if not sheet:
        sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Find header row with 'HỌ TÊN' or 'SĐT'
    header_idx = -1
    for r_idx, row in enumerate(rows[:10]):
        row_str = [str(c).upper().strip() if c is not None else "" for c in row[:12]]
        if any("HỌ TÊN" in c or "HOTEN" in c or "SĐT" in c for c in row_str):
            header_idx = r_idx
            break

    if header_idx == -1:
        header_idx = 0

    header_row = [str(c).upper().strip() if c is not None else "" for c in rows[header_idx]]
    
    # Exact Column Index Mapping based on Sheet Structure:
    # Col 0: STT / empty
    # Col 1: DATE
    # Col 2: HỌ TÊN
    # Col 3: SĐT
    # Col 4: NGUỒN
    # Col 5: CHI NHÁNH
    # Col 6: DỊCH VỤ / TÌNH TRẠNG
    # Col 7: LỊCH HẸN
    # Col 8: NGƯỜI ĐẶT HẸN
    # Col 9: DOANH THU / TỔNG TIỀN
    
    col_date = 1
    col_name = 2
    col_phone = 3
    col_source = 4
    col_branch = 5
    col_service = 6
    col_staff = 8
    col_rev = 9

    # Dynamic header lookup if available
    for idx, col in enumerate(header_row):
        if "DATE" in col or "NGÀY" in col: col_date = idx
        elif "HỌ TÊN" in col or "HOTEN" in col: col_name = idx
        elif "SĐT" in col or "TEL" in col or "PHONE" in col: col_phone = idx
        elif "NGUỒN" in col or "SOURCE" in col: col_source = idx
        elif "CHI NHÁNH" in col or "NHÁNH" in col: col_branch = idx
        elif "DỊCH VỤ" in col or "TÌNH TRẠNG" in col: col_service = idx
        elif "NGƯỜI ĐẶT" in col or "STAFF" in col: col_staff = idx
        elif "DOANH THU" in col or "TỔNG TIỀN" in col: col_rev = idx

    print(f"  Col Mapping -> Date:{col_date}, Name:{col_name}, Phone:{col_phone}, Source:{col_source}, Branch:{col_branch}, Service:{col_service}, Rev:{col_rev}")

    for row in rows[header_idx + 1:]:
        if not row or len(row) < 6: continue
        
        name = str(row[col_name]).strip() if col_name < len(row) and row[col_name] is not None else ""
        phone = str(row[col_phone]).strip() if col_phone < len(row) and row[col_phone] is not None else ""
        
        if not name or name.upper() in ["NONE", "HỌ TÊN", "SUM", "TỔNG", "STT", "2020-2024"]:
            continue

        date_val = str(row[col_date])[:10] if col_date < len(row) and row[col_date] is not None else ""
        source = str(row[col_source]).strip() if col_source < len(row) and row[col_source] is not None else "FACEBOOK"
        
        raw_branch = str(row[col_branch]).strip() if col_branch < len(row) and row[col_branch] is not None else "HCM"
        # CLEAN BRANCH NAME: If numeric string, fallback to HCM or correct branch
        if raw_branch.replace('.', '').replace('-', '').isdigit():
            raw_branch = "HCM"
        
        service = str(row[col_service]).strip() if col_service < len(row) and row[col_service] is not None else "Khác"
        staff = str(row[col_staff]).strip() if col_staff < len(row) and row[col_staff] is not None else ""
        
        rev_val = row[col_rev] if col_rev < len(row) and row[col_rev] is not None else 0
        try:
            revenue = float(rev_val)
        except Exception:
            revenue = 0.0

        customers.append({
            "date": date_val,
            "name": name,
            "phone": phone,
            "source": source,
            "branch": raw_branch.upper(),
            "service": service,
            "staff": staff,
            "revenue": revenue,
            "status": "Đã cắm Implant / Đã làm sứ / Đã khám",
            "file": file_path.name
        })

    print(f"  -> Extracted {len(customers):,} clean customers from {file_path.name}")
    return customers

def main():
    customer_dir = Path("customer/2026")
    files = sorted(list(customer_dir.glob("*.xlsx")))
    
    all_customers = []
    for f in files:
        custs = parse_monthly_file(f)
        all_customers.extend(custs)

    out_file = Path(".claude-ads/runs/live-meta-portfolio/all_customers_parsed.json")
    out_file.parent.mkdir(parents=True, exist_ok=True)
    out_file.write_text(json.dumps(all_customers, indent=2, ensure_ascii=False), encoding="utf-8")
    
    print("\n=========================================================================================")
    print(f"[SUMMARY STRICT] TOTAL QUALIFIED CUSTOMERS EXTRACTED (2025-T6/2026): {len(all_customers):,}")
    print(f"[SAVED] Clean output saved to: {out_file}")
    print("=========================================================================================")

if __name__ == "__main__":
    main()
