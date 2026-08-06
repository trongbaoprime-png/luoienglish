"""Parse sheet tab 'DATHEN' from data/google_leads_full.xlsx and extract all ~19,000 appointment leads."""

import sys
import openpyxl
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

def parse_dathen():
    xlsx_path = Path("data/google_leads_full.xlsx")
    if not xlsx_path.exists():
        print("[ERROR] data/google_leads_full.xlsx not found.")
        return

    print("[PARSER] Loading workbook data/google_leads_full.xlsx...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    print("\nWorkbook Sheet Names:")
    for idx, name in enumerate(wb.sheetnames):
        print(f"  {idx}: '{name}'")

    # Find DATHEN sheet
    dathen_sheet_name = None
    for name in wb.sheetnames:
        if "DATHEN" in name.upper() or "ĐÃ HẸN" in name.upper() or "DA HEN" in name.upper():
            dathen_sheet_name = name
            break

    if not dathen_sheet_name:
        dathen_sheet_name = wb.sheetnames[0]

    print(f"\n[TARGET SHEET TAB] Selected Sheet: '{dathen_sheet_name}'")
    sheet = wb[dathen_sheet_name]

    rows = []
    for r in sheet.iter_rows(values_only=True):
        if r and any(r):
            rows.append([str(cell).strip() if cell is not None else "" for cell in r])

    print(f"Total Non-Empty Rows Extracted from '{dathen_sheet_name}': {len(rows):,}")

    if len(rows) > 0:
        print("\nRow 0 (Header/Title):", rows[0][:12])
    if len(rows) > 1:
        print("Row 1 (Header/Columns):", rows[1][:12])
    if len(rows) > 2:
        print("Row 2 (Sample Data 1):", rows[2][:12])
    if len(rows) > 3:
        print("Row 3 (Sample Data 2):", rows[3][:12])

if __name__ == "__main__":
    parse_dathen()
