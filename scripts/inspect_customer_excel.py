"""Inspect Excel file structure, sheets, column headers, and data types for customer revenue files."""

import sys
import json
from pathlib import Path
import openpyxl

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

def inspect_file(file_path: Path):
    print(f"\n==================================================")
    print(f"INSPECTING FILE: {file_path.name}")
    print(f"==================================================")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheets in workbook ({len(sheet_names)}): {sheet_names}")

    # Inspect first active sheet
    sheet = wb[sheet_names[0]]
    rows = list(sheet.iter_rows(values_only=True, max_row=15))

    print(f"\n--- Top 15 rows of sheet '{sheet_names[0]}' ---")
    for r_idx, row in enumerate(rows, start=1):
        non_empty = [str(cell) for cell in row if cell is not None]
        if non_empty:
            print(f"Row {r_idx:2d}: {non_empty[:8]}")

def main():
    customer_dir = Path("customer/2026")
    excel_files = sorted(list(customer_dir.glob("*.xlsx")))

    print(f"Found {len(excel_files)} Excel files in {customer_dir.absolute()}")

    # Inspect recent 2026 file (T6.2026) and one 2025 file (T12.2025)
    sample_files = [
        customer_dir / "DT T6.2026.xlsx",
        customer_dir / "DT T1.2026.xlsx",
        customer_dir / "DT T12.2025.xlsx"
    ]

    for fpath in sample_files:
        if fpath.exists():
            inspect_file(fpath)

if __name__ == "__main__":
    main()
